#!/usr/bin/env python
"""Generate the Wayward China Commission recovery report (xlsx) from LIVE data.

Reads the commission engine (lens_ps_claim / lens_ps_commission_ledger, cip_104) and writes a
3-tab workbook — Invoice Now · Full List · How this was calculated — into this folder, stamped with
today's date. Re-run anytime; the numbers recompute from current data.

    # ambient DATABASE_URL must be the CIP prod DB (or a local override)
    .venv/Scripts/python.exe WORKBENCH/china-audit/reports/generate_recovery_report.py

Read-only against the DB. The .xlsx output is a generated artifact (gitignored); this script is the
source of record.
"""
from __future__ import annotations

import datetime
import os
from pathlib import Path

import psycopg
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PS_TENANT = "078a37d6-6ae2-4e22-869e-cc08f6cb2787"
MATERIAL_THRESHOLD = 5.0  # brands owed >= this are the "invoice now" set; below is rounding tail

OWN = {"never_listed": "Ours", "flat_fee_era_eric": "Flat-fee (ours from Dec)"}
PROD = {"boosted+connect": "Connect + Boost", "connect": "Connect", "boosted": "Boost", "": "—"}
HDR = ["#", "Brand", "Ownership", "Products", "Collected (since anchor)", "Fee owed (us)",
       "Wayward paid", "STILL OWED", "Referral partner"]
NAVY, LIGHTGOLD, GREY, HEAD = "1F3A5F", "FBF3D9", "8A8A8A", "2E4A6B"


def fetch_rows(conn: psycopg.Connection) -> list[tuple]:
    conn.execute("SELECT set_config('app.current_tenant', %s, false)", (PS_TENANT,))
    return conn.execute("""
        WITH coll AS (
            SELECT wayward_brand_id,
                   round(sum(usage_collected) FILTER (WHERE claimable), 2) AS collected,
                   string_agg(DISTINCT product_id, '+' ORDER BY product_id)
                       FILTER (WHERE claimable) AS products
            FROM lens_ps_commission_ledger GROUP BY 1)
        SELECT cl.brand_name, cl.ownership, COALESCE(co.products, ''),
               COALESCE(co.collected, 0), cl.mgmt_fee_owed, cl.wayward_paid, cl.ps_claim_owed,
               COALESCE(NULLIF(cl.partner_of_record, 'unassigned'), '')
        FROM lens_ps_claim cl
        LEFT JOIN coll co USING (wayward_brand_id)
        WHERE cl.verdict = 'china' AND cl.ps_claim_owed > 0
        ORDER BY cl.ps_claim_owed DESC
    """).fetchall()


def _row_cells(r: tuple) -> list:
    brand, own, prod, coll, fee, paid, owed, partner = r
    return [brand, OWN.get(own, own), PROD.get(prod, prod or "—"),
            float(coll), float(fee), float(paid), float(owed), partner]


def _write_list(ws, data: list[tuple], title: str, subtitle: str) -> None:
    thin = Side(style="thin", color="D9D9D9")
    border = Border(thin, thin, thin, thin)
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True, color=NAVY)
    ws["A2"] = subtitle
    ws["A2"].font = Font(size=10, italic=True, color=GREY)
    ws["A4"] = "STILL OWED"
    ws["A4"].font = Font(size=11, bold=True, color=NAVY)
    ws["C4"] = round(sum(float(r[6]) for r in data), 2)
    ws["C4"].number_format = "$#,##0.00"
    ws["C4"].font = Font(size=14, bold=True, color="B00020")
    ws["D4"] = f"{len(data)} brands"
    ws["D4"].font = Font(size=10, color=GREY)
    hr = 6
    for ci, h in enumerate(HDR, 1):
        c = ws.cell(hr, ci, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HEAD)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    for i, r in enumerate(data, 1):
        rr = hr + i
        for ci, v in enumerate([i, *_row_cells(r)], 1):
            c = ws.cell(rr, ci, v)
            c.border = border
            if ci in (5, 6, 7, 8):
                c.number_format = "$#,##0.00"
            if ci == 1:
                c.alignment = Alignment(horizontal="center")
        if float(r[6]) >= MATERIAL_THRESHOLD:
            for ci in range(1, 10):
                ws.cell(rr, ci).fill = PatternFill("solid", fgColor=LIGHTGOLD)
            ws.cell(rr, 8).font = Font(bold=True, color=NAVY)
        else:
            for ci in range(1, 10):
                ws.cell(rr, ci).font = Font(color=GREY)
    for ci, w in enumerate([4, 34, 22, 16, 20, 14, 14, 14, 16], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A7"
    ws.auto_filter.ref = f"A{hr}:I{hr + len(data)}"


def _write_methodology(ws) -> None:
    notes = [
        ("Wayward China Commission Recovery — how the numbers are built", 16, True, NAVY),
        ("", 10, False, GREY),
        ("WHAT THIS IS", 12, True, NAVY),
        ("Money Wayward has collected from Chinese brands on the Project Silk platform but has NOT "
         "paid us our management commission on. Computed live from the payment data.",
         11, False, "000000"),
        ("", 10, False, GREY),
        ("WHO COUNTS (a brand is 'ours' to claim when ALL are true)", 12, True, NAVY),
        ("1. It is Chinese  (verdict = china in our nationality audit).", 11, False, "000000"),
        ("2. It is ours — not on the partner-exclusion list (where a partner is already paid). The "
         "'Eric Flat Fee' bucket counts as ours, but only from Dec 2025.", 11, False, "000000"),
        ("3. Revenue is on/after our start date — Oct 1 2025 never-listed, Dec 1 2025 flat-fee.",
         11, False, "000000"),
        ("", 10, False, GREY),
        ("THE MATH (per brand, per product, per month)", 12, True, NAVY),
        ("Fee owed = usage the client PAID Wayward ('collected') x our rate. Ladder: 10% first 12 "
         "months, then 6% for 6 months, then 3% (restarts on a genuine win-back).",
         11, False, "000000"),
        ("Still owed = Fee owed − what Wayward already paid us (floored at $0 per brand).",
         11, False, "000000"),
        ("", 10, False, GREY),
        ("EXCLUDED ON PURPOSE", 12, True, NAVY),
        ("Non-Chinese brands; exclusion-list brands a partner is paid on; revenue before the start "
         "date; usage billed but not yet collected; pre-cutover referral commissions.",
         11, False, "000000"),
        ("", 10, False, GREY),
        ("CAVEATS", 12, True, NAVY),
        ("Data runs through the latest synced month; the number grows over time. Partner "
         "rates use a 5% placeholder pending Rhea's roster. A claim formally handed to "
         "Wayward is frozen as a dated snapshot so it cannot shift mid-negotiation.",
         11, False, "000000"),
    ]
    for i, (t, sz, b, col) in enumerate(notes, 1):
        c = ws.cell(i, 1, t)
        c.font = Font(size=sz, bold=b, color=col)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 120


def main() -> None:
    url = os.environ["DATABASE_URL"].replace("postgresql+psycopg://", "postgresql://")
    with psycopg.connect(url) as conn:
        rows = fetch_rows(conn)
    material = [r for r in rows if float(r[6]) >= MATERIAL_THRESHOLD]
    today = datetime.date.today().isoformat()

    wb = Workbook()
    _write_list(wb.active, material, "Wayward China Recovery — INVOICE NOW",
                f"Brands worth invoicing (owed >= ${MATERIAL_THRESHOLD:.0f}). As of {today}.")
    wb.active.title = "Invoice Now"
    _write_list(wb.create_sheet(f"Full List ({len(rows)})"), rows,
                "Wayward China Recovery — FULL LIST",
                "Every Chinese brand with a balance. Grey rows under the threshold are rounding "
                "tail.")
    _write_methodology(wb.create_sheet("How this was calculated"))

    out = Path(__file__).parent / f"Wayward-China-Recovery-{today}.xlsx"
    wb.save(out)
    total = sum(float(r[6]) for r in rows)
    mtot = sum(float(r[6]) for r in material)
    print(f"SAVED: {out}")
    print(f"Invoice Now: {len(material)} brands ${mtot:,.2f}  |  "
          f"Full: {len(rows)} brands ${total:,.2f}")


if __name__ == "__main__":
    main()
