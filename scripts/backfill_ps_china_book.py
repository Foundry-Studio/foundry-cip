# foundry: kind=script domain=client-intelligence-platform touches=storage
"""Backfill the PS China Book v2 schema (cip_38..cip_40) from held artifacts.

PS China Book Schema v2, Phase 1 §S6 (build spec: china-commission-audit/12-CC-SCHEMA-HANDOFF.md).
Idempotent + re-runnable: every write is ON CONFLICT DO NOTHING or an
IS-DISTINCT-FROM guarded UPDATE, so re-running loads nothing new.

Sources (paths are CLI args; defaults point at the venture-ecomlever artifacts):
  - May 2026 referral report .xlsx  → ps_payment_events (304) + cip_clients
                                       (wayward_brand_id, lifecycle_status) +
                                       ps_partner_credit referral queue
  - EXCLUSION-LIST-EXHIBIT-A.csv     → cip_clients.exhibit_a (~235)
  - MASTER-FIGHT-70.csv              → ps_claims #001 (draft) + lines
  - MASTER-FINDERS-FEE-314.csv       → ps_partner_credit + ps_attribution (finders_fee)
  - rule pass over cip_clients       → nationality_class first pass (name-signals)

Every backfilled surface writes a ps_annotations provenance row
(author=agent:claude_code, source_ref=<artifact>).

WARNING: reads DATABASE_URL. In foundry-cip the ambient DATABASE_URL is PROD —
pass --database-url or set it to a local scratch DB for dry runs.

Usage:
  DATABASE_URL=postgresql+psycopg://postgres:postgres@localhost:5544x/db \
      python scripts/backfill_ps_china_book.py --report <xlsx> --audit-dir <dir> [--dry-run]
"""
from __future__ import annotations

import argparse
import csv
import os
import re
import sys
from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path

from sqlalchemy import create_engine, text

PS_TENANT = "078a37d6-6ae2-4e22-869e-cc08f6cb2787"
_AUTHOR = "agent:claude_code"

_DEF_REPORT = (
    "c:/Users/Tim Jordan/code/venture-ecomlever/clients/wayward/"
    "data/referral-reports/2026-05-referral-report.xlsx"
)
_DEF_AUDIT = (
    "c:/Users/Tim Jordan/code/venture-ecomlever/clients/wayward/china-commission-audit"
)


def _norm(name: str | None) -> str:
    """doc-08 match convention: drop the ' - <deal type>' suffix, then
    lowercase alphanumeric-only. 'Tumella - Brand Deal' -> 'tumella'."""
    if not name:
        return ""
    base = name.split(" - ")[0]
    return re.sub(r"[^a-z0-9]", "", base.lower())


def _dec(v: object) -> Decimal:
    if v in (None, ""):
        return Decimal("0")
    return Decimal(str(v))


def _parse_ts(v: object):
    if v in (None, ""):
        return None
    if isinstance(v, datetime):
        return v if v.tzinfo else v.replace(tzinfo=UTC)
    s = str(v).strip().replace(" Z", "+00:00").replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(s)
    except ValueError:
        return datetime.fromisoformat(s[:19]).replace(tzinfo=UTC)


def _annotate(conn, entity_type, entity_id, note_type, body, source_ref) -> None:
    conn.execute(
        text(
            "INSERT INTO ps_annotations "
            "(tenant_id, entity_type, entity_id, note_type, body, author, source_ref) "
            "VALUES (:t,:et,:eid,:nt,:body,:auth,:sref) "
            "ON CONFLICT (tenant_id, entity_type, entity_id, note_type, source_ref) DO NOTHING"
        ),
        {"t": PS_TENANT, "et": entity_type, "eid": str(entity_id), "nt": note_type,
         "body": body, "auth": _AUTHOR, "sref": source_ref},
    )


def _load_norm_map(conn) -> dict[str, str]:
    """normalized cip_clients.name -> client id (PS tenant). First match wins on
    collision (1486 distinct of 1505)."""
    rows = conn.execute(text("SELECT id, name FROM cip_clients")).fetchall()
    out: dict[str, str] = {}
    for cid, name in rows:
        k = _norm(name)
        out.setdefault(k, str(cid))
    return out


def backfill_payment_events(conn, ws, source_ref) -> dict:
    ins = text(
        """
        INSERT INTO ps_payment_events (
            tenant_id, client_id, customer_id, wayward_brand_id, brand_name,
            payment_date, signup_date, stripe_invoice_ids, stripe_invoice_links,
            commission_fees_paid, usage_fees_paid, saas_fees_paid,
            cc_processing_fees_paid, total_amount_paid, rev_share_stated,
            months_from_signup, rev_share_start_date, days_since_start, source_ref
        ) VALUES (
            :t, :client_id, :cust, :wbid, :brand,
            :pdate, :signup, :inv_ids, :inv_links,
            :comm, :usage, :saas, :cc, :total, :stated,
            :mfs, :rss, :dss, :sref
        )
        ON CONFLICT (tenant_id, customer_id, payment_date, stripe_invoice_ids) DO NOTHING
        """
    )
    norm_map = _load_norm_map(conn)
    n = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[1] is None:  # BRAND_ID
            continue
        brand = r[2]
        pdate = r[3].date() if isinstance(r[3], datetime) else r[3]
        rss = _parse_ts(r[14])
        conn.execute(ins, {
            "t": PS_TENANT,
            "client_id": norm_map.get(_norm(brand)),
            "cust": r[0], "wbid": str(r[1]), "brand": brand,
            "pdate": pdate, "signup": _parse_ts(r[4]),
            "inv_ids": r[5], "inv_links": r[6],
            "comm": str(_dec(r[7])), "usage": str(_dec(r[8])), "saas": str(_dec(r[9])),
            "cc": str(_dec(r[10])), "total": str(_dec(r[11])), "stated": str(_dec(r[12])),
            "mfs": r[13], "rss": rss.date() if rss else None, "dss": r[15],
            "sref": source_ref,
        })
        n += 1
    _annotate(conn, "ps_payment_events", "batch", "provenance",
              f"{n} May-2026 payment events loaded from referral report", source_ref)
    return {"payment_events_processed": n}


def backfill_brand_list(conn, ws, source_ref) -> dict:
    upd = text(
        """
        UPDATE cip_clients
           SET wayward_brand_id = :wbid,
               lifecycle_status = :status
         WHERE id = :cid
           AND (wayward_brand_id IS DISTINCT FROM :wbid
                OR lifecycle_status IS DISTINCT FROM :status)
        """
    )
    credit = text(
        """
        INSERT INTO ps_partner_credit (tenant_id, client_id, referral_detail_raw)
        VALUES (:t, :cid, :raw)
        ON CONFLICT (tenant_id, client_id, product_id) DO UPDATE
           SET referral_detail_raw = EXCLUDED.referral_detail_raw
         WHERE ps_partner_credit.referral_detail_raw IS DISTINCT FROM EXCLUDED.referral_detail_raw
        """
    )
    norm_map = _load_norm_map(conn)
    matched = 0
    referral_q = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[1] is None:
            continue
        cid = norm_map.get(_norm(r[0]))
        if not cid:
            continue
        conn.execute(upd, {"cid": cid, "wbid": str(r[1]), "status": r[14]})
        matched += 1
        raw = r[4]  # Referral Source
        if raw:
            conn.execute(credit, {"t": PS_TENANT, "cid": cid, "raw": raw})
            referral_q += 1
    _annotate(conn, "cip_clients", "brand_list", "provenance",
              f"{matched} brands updated (wayward_brand_id + lifecycle_status)", source_ref)
    return {"brand_list_rows_matched": matched, "referral_credit_seeded": referral_q}


def backfill_exhibit_a(conn, rows, source_ref) -> dict:
    upd = text(
        """
        UPDATE cip_clients
           SET exhibit_a = true, exhibit_a_matched_name = :mname
         WHERE id = :cid AND exhibit_a IS DISTINCT FROM true
        """
    )
    # Exhibit-A brands are OTHER partners' books, so they are largely NOT on
    # Tim's Brand List (which is all-Tim-assigned) — their wayward_brand_id is
    # never set on cip_clients. Normalized-name is the only usable key.
    norm_map = _load_norm_map(conn)
    flagged = 0
    for row in rows:
        cid = norm_map.get(_norm(row["brand"]))
        if not cid:
            continue
        conn.execute(upd, {"cid": cid, "mname": row["brand"]})
        flagged += 1
    _annotate(conn, "cip_clients", "exhibit_a", "provenance",
              f"{flagged} brands flagged exhibit_a from contract list", source_ref)
    return {"exhibit_a_flagged": flagged}


def backfill_fight70(conn, rows, source_ref) -> dict:
    conn.execute(
        text(
            "INSERT INTO ps_claims (tenant_id, claim_number, claim_type, status) "
            "VALUES (:t, '001', 'uncredited_chinese', 'draft') "
            "ON CONFLICT (tenant_id, claim_number) DO NOTHING"
        ),
        {"t": PS_TENANT},
    )
    claim_id = conn.execute(
        text("SELECT id FROM ps_claims WHERE tenant_id=:t AND claim_number='001'"),
        {"t": PS_TENANT},
    ).scalar()
    line = text(
        """
        INSERT INTO ps_claim_lines (tenant_id, claim_id, client_id, amount, note, source_ref)
        VALUES (:t, :claim, :cid, :amt, :note, :sref)
        ON CONFLICT DO NOTHING
        """
    )
    # ps_claim_lines has no natural unique key; guard idempotency by only
    # inserting when this claim has no lines yet.
    existing = conn.execute(
        text("SELECT count(*) FROM ps_claim_lines WHERE tenant_id=:t AND claim_id=:c"),
        {"t": PS_TENANT, "c": claim_id},
    ).scalar()
    norm_map = _load_norm_map(conn)
    n = 0
    if not existing:
        for row in rows:
            cid = norm_map.get(_norm(row["brand"]))
            conn.execute(line, {
                "t": PS_TENANT, "claim": claim_id, "cid": cid,
                "amt": str(_dec(row["commission_10pct_since_dec1_usd"])),
                "note": (
                    f"{row['brand']} | tagged={row['currently_tagged']} "
                    f"| deal={row['hubspot_deal_id']}"
                ),
                "sref": source_ref,
            })
            n += 1
    _annotate(conn, "ps_claims", str(claim_id), "provenance",
              f"claim 001 seeded from Fight-70 ({n} lines)", source_ref)
    return {"claim_001_lines": n or existing}


def backfill_finders_fee(conn, rows, source_ref) -> dict:
    credit = text(
        """
        INSERT INTO ps_partner_credit (tenant_id, client_id, referral_detail_raw)
        VALUES (:t, :cid, :raw)
        ON CONFLICT (tenant_id, client_id, product_id) DO UPDATE
           SET referral_detail_raw = EXCLUDED.referral_detail_raw
         WHERE ps_partner_credit.referral_detail_raw IS DISTINCT FROM EXCLUDED.referral_detail_raw
        """
    )
    attr = text(
        """
        INSERT INTO ps_attribution
            (tenant_id, client_id, product_id, ps_conditional, ps_lead_source,
             changed_by, change_reason)
        SELECT :t, :cid, 'connect', 'finders_fee', :src, :author, 'finders-fee-314 backfill'
        WHERE NOT EXISTS (
            SELECT 1 FROM ps_attribution
             WHERE tenant_id=:t AND client_id=:cid AND product_id='connect'
               AND ps_conditional='finders_fee' AND effective_to IS NULL
        )
        """
    )
    norm_map = _load_norm_map(conn)
    credited = 0
    attributed = 0
    for row in rows:
        cid = norm_map.get(_norm(row["brand"]))
        if not cid:
            continue
        conn.execute(credit, {"t": PS_TENANT, "cid": cid, "raw": row["partner_source"]})
        credited += 1
        conn.execute(attr, {"t": PS_TENANT, "cid": cid, "src": row["partner_source"],
                            "author": _AUTHOR})
        attributed += 1
    _annotate(conn, "ps_attribution", "finders_fee_314", "provenance",
              f"{attributed} finders-fee conditional rows seeded", source_ref)
    return {"finders_fee_credited": credited, "finders_fee_attributed": attributed}


_CJK = re.compile(r"[一-鿿㐀-䶿]")


def classification_pass(conn, source_ref) -> dict:
    """Phase-1 first pass over cip_clients using name-only signals (CJK chars).
    Country/domain/email/phone signals need cip_companies/contacts joins — that
    is the Phase-2 recurring job (S8 out of scope). Confirmed only on strong
    (CJK-in-name) evidence; everything else left as-is ('unknown')."""
    rows = conn.execute(text("SELECT id, name FROM cip_clients")).fetchall()
    upd = text(
        "UPDATE cip_clients SET nationality_class='chinese_confirmed' "
        "WHERE id=:cid AND nationality_class <> 'chinese_confirmed'"
    )
    confirmed = 0
    for cid, name in rows:
        if name and _CJK.search(name):
            conn.execute(upd, {"cid": str(cid)})
            confirmed += 1
    _annotate(conn, "cip_clients", "classification", "provenance",
              f"name-signal classification pass: {confirmed} chinese_confirmed (CJK-in-name)",
              source_ref)
    return {"classification_confirmed": confirmed, "clients_scanned": len(rows)}


def _read_csv(path: Path) -> list[dict]:
    with path.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def run_backfill(engine, report_path: Path, audit_dir: Path, *, dry_run: bool) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(report_path, data_only=True)
    report_ref = report_path.name
    summary: dict = {}
    with engine.begin() as conn:
        conn.execute(text("SELECT set_config('app.current_tenant', :t, false)"),
                     {"t": PS_TENANT})
        summary |= backfill_payment_events(conn, wb["Rev Share Report"], report_ref)
        summary |= backfill_brand_list(conn, wb["Brand List"], report_ref)
        summary |= backfill_exhibit_a(
            conn, _read_csv(audit_dir / "EXCLUSION-LIST-EXHIBIT-A.csv"),
            "EXCLUSION-LIST-EXHIBIT-A.csv")
        summary |= backfill_fight70(
            conn, _read_csv(audit_dir / "MASTER-FIGHT-70.csv"), "MASTER-FIGHT-70.csv")
        summary |= backfill_finders_fee(
            conn, _read_csv(audit_dir / "MASTER-FINDERS-FEE-314.csv"),
            "MASTER-FINDERS-FEE-314.csv")
        summary |= classification_pass(conn, "ps_classification_rules")
        # Verification (read inside the txn)
        pe = conn.execute(text(
            "SELECT count(*), coalesce(sum(usage_fees_paid),0), "
            "coalesce(sum(rev_share_stated),0), count(*) FILTER (WHERE rev_share_variance <> 0) "
            "FROM ps_payment_events WHERE tenant_id=:t"), {"t": PS_TENANT}).fetchone()
        summary["verify_payment_events"] = {
            "rows": pe[0], "sum_usage_fees_paid": str(pe[1]),
            "sum_rev_share_stated": str(pe[2]), "rows_with_variance": pe[3],
        }
        if dry_run:
            conn.execute(text("ROLLBACK"))
            summary["DRY_RUN"] = "rolled back"
    return summary


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--report", default=_DEF_REPORT)
    ap.add_argument("--audit-dir", default=_DEF_AUDIT)
    ap.add_argument("--database-url", default=None)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args(argv)
    url = args.database_url or os.environ.get("DATABASE_URL")
    if not url:
        print("DATABASE_URL not set", file=sys.stderr)
        return 2
    if url.startswith("postgresql://"):
        url = url.replace("postgresql://", "postgresql+psycopg://", 1)
    engine = create_engine(url, pool_pre_ping=True)
    try:
        summary = run_backfill(engine, Path(args.report), Path(args.audit_dir),
                               dry_run=args.dry_run)
    finally:
        engine.dispose()
    import json
    print(json.dumps(summary, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
