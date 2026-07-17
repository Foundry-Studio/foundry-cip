# foundry: kind=script domain=client-intelligence-platform touches=storage
"""Ingest Jake's monthly rev-share reports (all months) into ps_payment_events.

SCHEMA DRIFT is the whole problem this solves. Jake's report layout is NOT stable:
May 2026 carries `Rev Share Start Date` + `Days Since Rev Share Start`; June 2026
instead carries `MONTHS_FROM_SIGNUP_TO_PAYMENT_1`. A loader that assumes fixed
columns silently produces wrong money.

How drift is handled (three rules, in order of importance):

  1. MAP BY HEADER NAME, never by position. Each canonical field has an alias set.
  2. REQUIRED columns must be present, or the file is REJECTED. We never guess at a
     missing money column — a wrong number is worse than no number.
  3. UNKNOWN columns are reported, not ignored. New columns are how you find out Jake
     started sending something new (e.g. a per-product fee split — the D3 ask). Silent
     tolerance would hide that.

Plus a cross-check that catches both parser bugs AND Wayward's own arithmetic:
  4. Σ REV_SHARE_OWED in the file must equal the "Total Owed" Jake stated in the email.
     Those stated totals live in a SIDECAR CSV, not in this code (see below). A mismatch
     is reported loudly and, unless --force, the month is not committed.

The expected totals — the sidecar (why it's a CSV, not a dict):
  Jake's "Total Owed" per month used to be a hardcoded dict here, so every new month was a
  Python edit. It now lives next to the reports as EXPECTED-TOTALS.csv — a plain CSV Tim
  edits (header: month,expected_total,source_note; month = YYYY-MM). Default path is
  <reports-dir>/EXPECTED-TOTALS.csv; --expected-totals <path> overrides.
  5. A report month ABSENT from the sidecar is a LOUD REJECT of that month (unless
     --force) — exactly like a missing required column. Silence about a month is how a
     month gets forgotten; we refuse to load one we were never told the total for. Add its
     row to the sidecar, or pass --force to load it without the cross-check.

Idempotent: re-running loads nothing new (ON CONFLICT DO NOTHING on the natural key).

Usage:
  DATABASE_URL=... python scripts/ingest_payment_reports.py \
      [--dir <reports>] [--expected-totals <csv>] [--dry-run] [--force]
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path

from sqlalchemy import create_engine, text

PS_TENANT = "078a37d6-6ae2-4e22-869e-cc08f6cb2787"
_DEF_DIR = (
    "c:/Users/Tim Jordan/code/venture-ecomlever/clients/wayward/data/referral-reports"
)

# The totals Jake stated in his emails now live in a SIDECAR CSV next to the reports, not
# in this file — see load_expected_totals() and the module docstring. This is the default
# filename looked for inside --dir; --expected-totals overrides the whole path.
_EXPECTED_TOTALS_FILENAME = "EXPECTED-TOTALS.csv"

# canonical field -> accepted header aliases (lowercased, non-alnum stripped)
_ALIASES: dict[str, tuple[str, ...]] = {
    "customer_id": ("customerid",),
    "wayward_brand_id": ("brandid",),
    "brand_name": ("brandname",),
    "payment_date": ("paymentdate",),
    "signup_date": ("signupdate",),
    "stripe_invoice_ids": ("stripeinvoiceids",),
    "stripe_invoice_links": ("stripeinvoicelinks",),
    "commission_fees_paid": ("commissionfeespaid",),
    "usage_fees_paid": ("usagefeespaid",),
    "saas_fees_paid": ("saasfeespaid",),
    "cc_processing_fees_paid": ("ccprocessingfeespaid",),
    "total_amount_paid": ("totalamountpaid",),
    "rev_share_stated": ("revshareowed",),
    "months_from_signup": ("monthsfromsignuptopayment",),
    "rev_share_start_date": ("revsharestartdate",),
    "days_since_start": (
        "dayssincerevsharestart",
        "daysfromsignuptopayment",
        # The SAME slot, three names across three months. June's Google export
        # auto-suffixed the duplicate header; January's raw file just repeats it
        # verbatim (we suffix it ourselves in _dedupe). Values are days (e.g. 366),
        # not months — the header name lies.
        "monthsfromsignuptopayment1",
    ),
}

# Without these, the money is not trustworthy. Reject the file.
_REQUIRED = (
    "wayward_brand_id",
    "payment_date",
    "usage_fees_paid",
    "rev_share_stated",
)


def _norm_header(h: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (h or "").lower())


def _dec(v) -> Decimal:
    if v in (None, ""):
        return Decimal("0")
    s = str(v).replace("$", "").replace(",", "").strip()
    if not s or s in {"-", "NA", "N/A"}:
        return Decimal("0")
    try:
        return Decimal(s)
    except Exception:
        return Decimal("0")


def load_expected_totals(path: Path) -> dict[str, Decimal]:
    """Read the EXPECTED-TOTALS sidecar CSV -> {"YYYY-MM": Decimal}.

    Header: month,expected_total,source_note (source_note is documentation; ignored here).
    Columns are matched by normalized NAME (like the report loader), so case/spacing drift
    doesn't break it, and a month value is truncated to YYYY-MM so a full date is tolerated.
    A MISSING file returns {} — which makes every month reject (unless --force): a lost
    sidecar fails loud, never silent. A file that HAS rows but LACKS the required headers is
    a hard error (a broken sidecar must not masquerade as an empty one)."""
    totals: dict[str, Decimal] = {}
    if not path.exists():
        return totals
    with path.open(encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    if not rows:
        return totals
    header = [_norm_header(h) for h in rows[0]]
    try:
        mi = header.index("month")
        ti = header.index("expectedtotal")
    except ValueError as exc:
        raise SystemExit(
            f"{path}: EXPECTED-TOTALS sidecar needs headers 'month,expected_total,source_note'"
        ) from exc
    for row in rows[1:]:
        if len(row) <= max(mi, ti):
            continue
        month = (row[mi] or "").strip()[:7]
        if month:
            totals[month] = _dec(row[ti])
    return totals


def _date(v):
    if not v:
        return None
    s = str(v).strip()
    for f in ("%Y-%m-%d", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:10], f).date()
        except ValueError:
            continue
    return None


def _ts(v):
    if not v:
        return None
    s = str(v).strip().replace(" Z", "+00:00").replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(s)
    except ValueError:
        d = _date(s)
        return datetime(d.year, d.month, d.day, tzinfo=UTC) if d else None


def _dedupe(headers: list[str]) -> list[str]:
    """Jake's raw January file emits MONTHS_FROM_SIGNUP_TO_PAYMENT TWICE — the second
    is actually DAYS. dict(zip(headers, row)) would silently let days overwrite months.
    Suffix repeats so each column keeps its own value (Google's own CSV export does the
    same, producing ..._1)."""
    seen: dict[str, int] = {}
    out = []
    for h in headers:
        key = (h or "").strip()
        if key in seen:
            seen[key] += 1
            out.append(f"{key}_{seen[key]}")
        else:
            seen[key] = 0
            out.append(key)
    return out


def _rows_from(path: Path) -> tuple[list[str], list[dict]]:
    if path.suffix.lower() == ".csv":
        with path.open(encoding="utf-8-sig", newline="") as f:
            r = csv.reader(f)
            headers = _dedupe(next(r))
            return headers, [dict(zip(headers, row, strict=False)) for row in r]
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    headers = _dedupe([str(h) if h is not None else "" for h in next(it)])
    out = []
    for row in it:
        out.append(dict(zip(headers, row, strict=False)))
    return headers, out


def _build_map(headers: list[str]) -> tuple[dict[str, str], list[str]]:
    """canonical -> actual header. Also return UNKNOWN headers (drift signal)."""
    norm_to_actual = {}
    for h in headers:
        n = _norm_header(h)
        if n and n not in norm_to_actual:
            norm_to_actual[n] = h
    field_map: dict[str, str] = {}
    claimed: set[str] = set()
    for field, aliases in _ALIASES.items():
        for a in aliases:
            if a in norm_to_actual:
                field_map[field] = norm_to_actual[a]
                claimed.add(a)
                break
    # Blank trailing columns (and the _1/_2 suffixes _dedupe gives them) are not drift.
    # A drift alarm that cries wolf gets ignored, which defeats the point of having one.
    unknown = [
        h for h in headers
        if _norm_header(h) and _norm_header(h) not in claimed
        and not re.fullmatch(r"_?\d+", _norm_header(h))
    ]
    return field_map, unknown


_INSERT = text(
    """
    INSERT INTO ps_payment_events (
        tenant_id, client_id, customer_id, wayward_brand_id, brand_name,
        payment_date, signup_date, stripe_invoice_ids, stripe_invoice_links,
        commission_fees_paid, usage_fees_paid, saas_fees_paid,
        cc_processing_fees_paid, total_amount_paid, rev_share_stated,
        months_from_signup, rev_share_start_date, days_since_start, source_ref
    ) VALUES (
        :t, :cid, :cust, :wbid, :brand, :pdate, :signup, :inv, :links,
        :comm, :usage, :saas, :cc, :total, :stated, :mfs, :rss, :dss, :sref
    )
    ON CONFLICT (tenant_id, customer_id, payment_date, stripe_invoice_ids) DO NOTHING
    """
)


def ingest_file(
    conn, path: Path, wbid_map: dict[str, str],
    expected_totals: dict[str, Decimal], *, force: bool,
) -> dict:
    month = path.stem[:7]
    headers, rows = _rows_from(path)
    field_map, unknown = _build_map(headers)

    missing = [f for f in _REQUIRED if f not in field_map]
    if missing:
        return {"file": path.name, "REJECTED": f"missing required columns: {missing}",
                "headers_seen": headers}

    # A month we were never told the total for is a month we refuse to load (unless --force),
    # exactly like a missing required column — silence about a month is how one gets forgotten.
    if month not in expected_totals and not force:
        return {"file": path.name, "month": month,
                "REJECTED": f"month {month} not in EXPECTED-TOTALS sidecar — add a "
                            "'month,expected_total,source_note' row (or use --force)",
                "unknown_columns (DRIFT)": unknown}

    def g(row, field):
        h = field_map.get(field)
        return row.get(h) if h else None

    parsed, total_stated = 0, Decimal("0")
    payload = []
    for row in rows:
        wbid = (g(row, "wayward_brand_id") or "").strip() if g(row, "wayward_brand_id") else ""
        pdate = _date(g(row, "payment_date"))
        if not wbid or not pdate:
            continue
        stated = _dec(g(row, "rev_share_stated"))
        total_stated += stated
        parsed += 1
        rss = _ts(g(row, "rev_share_start_date"))
        payload.append({
            "t": PS_TENANT, "cid": wbid_map.get(wbid),
            "cust": g(row, "customer_id"), "wbid": wbid,
            "brand": g(row, "brand_name"), "pdate": pdate,
            "signup": _ts(g(row, "signup_date")),
            "inv": g(row, "stripe_invoice_ids"),
            "links": g(row, "stripe_invoice_links"),
            "comm": str(_dec(g(row, "commission_fees_paid"))),
            "usage": str(_dec(g(row, "usage_fees_paid"))),
            "saas": str(_dec(g(row, "saas_fees_paid"))),
            "cc": str(_dec(g(row, "cc_processing_fees_paid"))),
            "total": str(_dec(g(row, "total_amount_paid"))),
            "stated": str(stated),
            "mfs": _int(g(row, "months_from_signup")),
            "rss": rss.date() if rss else None,
            "dss": _int(g(row, "days_since_start")),
            "sref": path.name,
        })

    # Cross-check against what Jake said in the email.
    expected = expected_totals.get(month)
    delta = (total_stated - expected) if expected is not None else None
    ok = expected is None or abs(delta) < Decimal("0.02")
    result = {
        "file": path.name, "month": month, "rows_parsed": parsed,
        "sum_rev_share": str(total_stated),
        "email_total_owed": str(expected) if expected else None,
        "matches_email": ok,
        "delta": str(delta) if delta is not None else None,
        "unknown_columns (DRIFT)": unknown,
    }
    if not ok and not force:
        result["SKIPPED"] = "sum does not match Jake's stated total; use --force to load anyway"
        return result

    for i in range(0, len(payload), 500):
        conn.execute(_INSERT, payload[i:i + 500])
    result["loaded"] = len(payload)
    return result


def _int(v):
    try:
        return int(float(str(v).strip()))
    except Exception:
        return None


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dir", default=_DEF_DIR)
    ap.add_argument("--expected-totals", default=None,
                    help="path to the EXPECTED-TOTALS sidecar CSV "
                         "(default: <dir>/EXPECTED-TOTALS.csv)")
    ap.add_argument("--database-url", default=None)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--force", action="store_true")
    args = ap.parse_args(argv)

    url = args.database_url or os.environ.get("DATABASE_URL")
    if not url:
        print("DATABASE_URL not set", file=sys.stderr)
        return 2
    if url.startswith("postgresql://"):
        url = url.replace("postgresql://", "postgresql+psycopg://", 1)

    expected_path = (
        Path(args.expected_totals) if args.expected_totals
        else Path(args.dir) / _EXPECTED_TOTALS_FILENAME
    )
    expected_totals = load_expected_totals(expected_path)
    if not expected_path.exists():
        print(f"WARNING: EXPECTED-TOTALS sidecar not found at {expected_path} — every month "
              "will REJECT unless --force. Create it (month,expected_total,source_note).",
              file=sys.stderr)

    # The sidecar itself is a .csv in --dir; never ingest it as a report.
    sidecar_names = {_EXPECTED_TOTALS_FILENAME, expected_path.name}
    files = sorted(
        p for p in Path(args.dir).iterdir()
        if p.suffix.lower() in {".csv", ".xlsx"} and not p.name.startswith("~")
        and p.name not in sidecar_names
    )
    engine = create_engine(url, pool_pre_ping=True)
    out = []
    try:
        with engine.begin() as conn:
            conn.execute(text("SELECT set_config('app.current_tenant', :t, false)"),
                         {"t": PS_TENANT})
            wbid_map = {
                str(w): str(c) for c, w in conn.execute(text(
                    "SELECT id, wayward_brand_id FROM cip_clients "
                    "WHERE wayward_brand_id IS NOT NULL")).fetchall()
            }
            for p in files:
                out.append(ingest_file(conn, p, wbid_map, expected_totals, force=args.force))
            summary = conn.execute(text(
                "SELECT count(*), coalesce(sum(rev_share_stated),0), "
                "count(DISTINCT source_ref) FROM ps_payment_events WHERE tenant_id=:t"
            ), {"t": PS_TENANT}).fetchone()
            if args.dry_run:
                conn.execute(text("ROLLBACK"))
    finally:
        engine.dispose()
    print(json.dumps({
        "files": out,
        "TOTAL_in_db": {
            "payment_events": summary[0],
            "sum_rev_share": str(summary[1]),
            "distinct_reports": summary[2],
        },
        "dry_run": args.dry_run,
    }, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
